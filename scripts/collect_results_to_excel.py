"""
Collect experiment results from test logs and export to Excel.

Parses metrics_log.txt / test_log.txt from all Table 2/3/4 experiments
and the Full Model, then generates one Excel file per table with
per-modality metrics (aAcc, mIoU, mAcc, mDice, mFscore, mPrecision, mRecall)
plus per-class breakdowns.

Usage:
    python scripts/collect_results_to_excel.py
    python scripts/collect_results_to_excel.py --work-root work_dirs/paper_experiments
    python scripts/collect_results_to_excel.py --tables table2 table3 table4

Output:
    work_dirs/paper_experiments/Table2_results.xlsx
    work_dirs/paper_experiments/Table3_results.xlsx
    work_dirs/paper_experiments/Table4_results.xlsx
"""

import argparse
import os
import re
import glob
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ============================================================================
# Log Parsing
# ============================================================================

def parse_summary_metrics(log_text):
    """Parse summary metrics from mmengine log output.

    Looks for lines like:
        04/02 02:10:00 - mmengine - INFO - Iter(test) ... aAcc: 82.45  mIoU: 65.32 ...
    or the final dict output.
    """
    metrics = {}

    # Pattern 1: mmengine log line with metrics
    # e.g., "aAcc: 82.45  mIoU: 65.32  mAcc: 72.18  mDice: 75.00  mFscore: 73.50  mPrecision: 71.20  mRecall: 75.80"
    metric_pattern = r'(aAcc|mIoU|mAcc|mDice|mFscore|mPrecision|mRecall)\s*:\s*([\d.]+)'
    matches = re.findall(metric_pattern, log_text)
    for key, val in matches:
        # Take the last occurrence (in case of multiple test runs)
        metrics[key] = float(val)

    return metrics


def parse_per_class_table(log_text):
    """Parse per-class PrettyTable from log.

    Looks for:
        per class results:
        +-------+-------+-------+...
        | Class | IoU   | Acc   |...
        +-------+-------+-------+...
        | cls1  | 45.23 | 67.89 |...
        | cls2  | 52.11 | 71.34 |...
        +-------+-------+-------+...
    """
    per_class = {}

    # Find the table after "per class results:"
    marker = 'per class results:'
    idx = log_text.rfind(marker)  # Use last occurrence
    if idx == -1:
        return per_class

    table_text = log_text[idx:]
    lines = table_text.split('\n')

    header = None
    for line in lines:
        line = line.strip()
        if not line or line.startswith('+'):
            continue
        if line.startswith('|'):
            cells = [c.strip() for c in line.split('|') if c.strip()]
            if header is None:
                header = cells
            else:
                if len(cells) == len(header):
                    class_name = cells[0]
                    per_class[class_name] = {}
                    for i, h in enumerate(header[1:], 1):
                        try:
                            per_class[class_name][h] = float(cells[i])
                        except ValueError:
                            per_class[class_name][h] = cells[i]

    return per_class


def parse_log_file(log_path):
    """Parse a single log file and return metrics dict."""
    if not os.path.exists(log_path):
        return None, None

    with open(log_path, 'r') as f:
        text = f.read()

    summary = parse_summary_metrics(text)
    per_class = parse_per_class_table(text)

    return summary, per_class


def find_log_file(work_dir, modal, prefer_metrics=True):
    """Find the log file for a given experiment and modality."""
    candidates = []
    if prefer_metrics:
        candidates.append(os.path.join(work_dir, f'metrics_{modal}', 'metrics_log.txt'))
    candidates.append(os.path.join(work_dir, f'test_{modal}', 'test_log.txt'))
    # Also try metrics log as fallback
    candidates.append(os.path.join(work_dir, f'metrics_{modal}', 'metrics_log.txt'))

    for path in candidates:
        if os.path.exists(path):
            return path

    # Last resort: search for any log in the directory
    for subdir_pattern in [f'metrics_{modal}', f'test_{modal}']:
        subdir = os.path.join(work_dir, subdir_pattern)
        if os.path.isdir(subdir):
            # Try to find mmengine log files
            log_files = glob.glob(os.path.join(subdir, '*.log'))
            if log_files:
                return sorted(log_files)[-1]

    return None


# ============================================================================
# Experiment Definitions
# ============================================================================

TABLE2_EXPERIMENTS = OrderedDict([
    ('Full Model', {
        'work_dir': 'work_dirs/floodnet/SwinmoeB/655',
        'modals': ['sar', 'rgb', 'GF'],
        'desc': 'Swin-B + MoE (E=8, K=3)',
    }),
    ('w/o MoE', {
        'work_dir': 'work_dirs/paper_experiments/table2/no_moe',
        'modals': ['sar', 'rgb', 'GF'],
        'desc': 'Remove MoE, use standard FFN',
    }),
    ('w/o ModalSpecificStem', {
        'work_dir': 'work_dirs/paper_experiments/table2/no_modal_specific_stem',
        'modals': ['sar', 'rgb', 'GF'],
        'desc': 'Shared patch embedding',
    }),
    ('w/o Modal Bias', {
        'work_dir': 'work_dirs/paper_experiments/table2/no_modal_bias',
        'modals': ['sar', 'rgb', 'GF'],
        'desc': 'No modal bias in gating',
    }),
    ('w/o Shared Experts', {
        'work_dir': 'work_dirs/paper_experiments/table2/no_shared_experts',
        'modals': ['sar', 'rgb', 'GF'],
        'desc': 'No shared experts',
    }),
    ('w/o Separate Decoder', {
        'work_dir': 'work_dirs/paper_experiments/table2/shared_decoder',
        'modals': ['sar', 'rgb', 'GF'],
        'desc': 'Shared decoder head',
    }),
])

TABLE3_EXPERIMENTS = OrderedDict([
    ('E=6, K=1', {
        'work_dir': 'work_dirs/paper_experiments/table3/e6_k1',
        'modals': ['sar', 'rgb', 'GF'],
    }),
    ('E=6, K=2', {
        'work_dir': 'work_dirs/paper_experiments/table3/e6_k2',
        'modals': ['sar', 'rgb', 'GF'],
    }),
    ('E=6, K=3', {
        'work_dir': 'work_dirs/paper_experiments/table3/e6_k3',
        'modals': ['sar', 'rgb', 'GF'],
    }),
    ('E=8, K=1', {
        'work_dir': 'work_dirs/paper_experiments/table3/e8_k1',
        'modals': ['sar', 'rgb', 'GF'],
    }),
    ('E=8, K=2', {
        'work_dir': 'work_dirs/paper_experiments/table3/e8_k2',
        'modals': ['sar', 'rgb', 'GF'],
    }),
    ('E=8, K=3 (Full)', {
        'work_dir': 'work_dirs/floodnet/SwinmoeB/655',
        'modals': ['sar', 'rgb', 'GF'],
    }),
])

TABLE4_EXPERIMENTS = OrderedDict([
    ('SAR-only', {
        'work_dir': 'work_dirs/paper_experiments/table4/sar_only',
        'modals': ['sar'],
    }),
    ('RGB-only', {
        'work_dir': 'work_dirs/paper_experiments/table4/rgb_only',
        'modals': ['rgb'],
    }),
    ('GF-only', {
        'work_dir': 'work_dirs/paper_experiments/table4/gf_only',
        'modals': ['GF'],
    }),
    ('Multi-Modal (Full)', {
        'work_dir': 'work_dirs/floodnet/SwinmoeB/655',
        'modals': ['sar', 'rgb', 'GF'],
    }),
])

SUMMARY_METRICS = ['aAcc', 'mIoU', 'mAcc', 'mDice', 'mFscore', 'mPrecision', 'mRecall']
MODAL_DISPLAY = {'sar': 'SAR', 'rgb': 'RGB', 'GF': 'GaoFen'}


# ============================================================================
# Excel Generation
# ============================================================================

def collect_table_data(experiments, work_root=''):
    """Collect all metrics for a table's experiments."""
    results = OrderedDict()

    for exp_name, exp_info in experiments.items():
        work_dir = exp_info['work_dir']
        if work_root and not os.path.isabs(work_dir):
            # Don't prepend work_root if work_dir is already an absolute-like path
            # that doesn't start with work_root
            if not work_dir.startswith('work_dirs/paper_experiments'):
                pass  # Keep original path (e.g., Full Model)
            # work_dir stays as is

        results[exp_name] = {}
        for modal in exp_info['modals']:
            log_path = find_log_file(work_dir, modal)
            if log_path:
                summary, per_class = parse_log_file(log_path)
                results[exp_name][modal] = {
                    'summary': summary or {},
                    'per_class': per_class or {},
                    'log_path': log_path,
                }
                found_metrics = list((summary or {}).keys())
                print(f"  [OK] {exp_name} / {MODAL_DISPLAY.get(modal, modal)}: "
                      f"{log_path} ({len(found_metrics)} metrics)")
            else:
                results[exp_name][modal] = {
                    'summary': {},
                    'per_class': {},
                    'log_path': None,
                }
                print(f"  [--] {exp_name} / {MODAL_DISPLAY.get(modal, modal)}: "
                      f"no log found in {work_dir}")

    return results


def write_excel(results, experiments, output_path, table_name):
    """Write results to Excel with openpyxl for better formatting."""
    wb = openpyxl.Workbook()

    # ---- Style definitions ----
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    modal_fill = {
        'sar': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'rgb': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
        'GF':  PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    }
    exp_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    best_font = Font(bold=True, color='C00000')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # ==================== Sheet 1: Summary Table ====================
    ws = wb.active
    ws.title = 'Summary'

    # Determine all modals used
    all_modals = []
    for exp_info in experiments.values():
        for m in exp_info['modals']:
            if m not in all_modals:
                all_modals.append(m)

    # Header row 1: Experiment | SAR (spanning) | RGB (spanning) | GF (spanning)
    row = 1
    ws.cell(row=row, column=1, value='Experiment').font = header_font_white
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).border = thin_border

    col = 2
    for modal in all_modals:
        modal_name = MODAL_DISPLAY.get(modal, modal)
        start_col = col
        for metric in SUMMARY_METRICS:
            ws.cell(row=row + 1, column=col, value=metric).font = header_font
            ws.cell(row=row + 1, column=col).fill = modal_fill.get(modal, exp_fill)
            ws.cell(row=row + 1, column=col).border = thin_border
            ws.cell(row=row + 1, column=col).alignment = center_align
            col += 1
        end_col = col - 1
        # Merge header for modal name
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col, value=modal_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)

    # Data rows
    data_row = row + 2
    # Track best values per (modal, metric) for highlighting
    metric_values = {(m, met): [] for m in all_modals for met in SUMMARY_METRICS}

    for exp_name in results:
        for modal in all_modals:
            if modal in results[exp_name]:
                summary = results[exp_name][modal]['summary']
                for met in SUMMARY_METRICS:
                    val = summary.get(met)
                    if val is not None:
                        metric_values[(modal, met)].append((exp_name, val))

    for exp_name, exp_data in results.items():
        ws.cell(row=data_row, column=1, value=exp_name).font = Font(bold=True)
        ws.cell(row=data_row, column=1).border = thin_border

        col = 2
        for modal in all_modals:
            if modal in exp_data and exp_data[modal]['summary']:
                summary = exp_data[modal]['summary']
                for metric in SUMMARY_METRICS:
                    val = summary.get(metric)
                    cell = ws.cell(row=data_row, column=col)
                    if val is not None:
                        cell.value = val
                        cell.number_format = '0.00'
                        # Bold best value
                        best_vals = metric_values.get((modal, metric), [])
                        if best_vals:
                            best_val = max(v for _, v in best_vals)
                            if val == best_val and len(best_vals) > 1:
                                cell.font = best_font
                    else:
                        cell.value = '-'
                    cell.border = thin_border
                    cell.alignment = center_align
                    col += 1
            else:
                for _ in SUMMARY_METRICS:
                    cell = ws.cell(row=data_row, column=col, value='-')
                    cell.border = thin_border
                    cell.alignment = center_align
                    col += 1

        data_row += 1

    # Auto-width
    for col_idx in range(1, col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.column_dimensions['A'].width = 25

    # ==================== Sheet 2: Per-Class Details ====================
    ws2 = wb.create_sheet('Per-Class Details')
    row = 1

    for exp_name, exp_data in results.items():
        for modal in all_modals:
            if modal not in exp_data or not exp_data[modal]['per_class']:
                continue

            per_class = exp_data[modal]['per_class']
            modal_name = MODAL_DISPLAY.get(modal, modal)

            # Section header
            ws2.cell(row=row, column=1,
                     value=f'{exp_name} - {modal_name}').font = Font(bold=True, size=12)
            ws2.cell(row=row, column=1).fill = modal_fill.get(modal, exp_fill)
            row += 1

            # Get metric columns from first class
            first_class = list(per_class.keys())[0]
            metric_cols = list(per_class[first_class].keys())

            # Header
            ws2.cell(row=row, column=1, value='Class').font = header_font
            ws2.cell(row=row, column=1).border = thin_border
            for j, met in enumerate(metric_cols, 2):
                ws2.cell(row=row, column=j, value=met).font = header_font
                ws2.cell(row=row, column=j).border = thin_border
                ws2.cell(row=row, column=j).alignment = center_align
            row += 1

            # Data
            for cls_name, cls_metrics in per_class.items():
                ws2.cell(row=row, column=1, value=cls_name).border = thin_border
                for j, met in enumerate(metric_cols, 2):
                    cell = ws2.cell(row=row, column=j)
                    cell.value = cls_metrics.get(met, '-')
                    cell.border = thin_border
                    cell.alignment = center_align
                    if isinstance(cell.value, float):
                        cell.number_format = '0.00'
                row += 1

            row += 1  # Blank row between sections

    # Auto-width sheet 2
    ws2.column_dimensions['A'].width = 20
    for col_idx in range(2, 15):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 13

    # Save
    wb.save(output_path)
    print(f'\n  => Saved: {output_path}')


def write_csv_fallback(results, experiments, output_path, table_name):
    """Fallback: write CSV if openpyxl not available."""
    import csv

    all_modals = []
    for exp_info in experiments.values():
        for m in exp_info['modals']:
            if m not in all_modals:
                all_modals.append(m)

    csv_path = output_path.replace('.xlsx', '.csv')

    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)

        # Header
        header = ['Experiment']
        for modal in all_modals:
            modal_name = MODAL_DISPLAY.get(modal, modal)
            for metric in SUMMARY_METRICS:
                header.append(f'{modal_name}_{metric}')
        writer.writerow(header)

        # Data
        for exp_name, exp_data in results.items():
            row = [exp_name]
            for modal in all_modals:
                if modal in exp_data and exp_data[modal]['summary']:
                    summary = exp_data[modal]['summary']
                    for metric in SUMMARY_METRICS:
                        val = summary.get(metric)
                        row.append(f'{val:.2f}' if val is not None else '-')
                else:
                    row.extend(['-'] * len(SUMMARY_METRICS))
            writer.writerow(row)

    print(f'\n  => Saved (CSV fallback): {csv_path}')


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Collect experiment results into Excel tables')
    parser.add_argument('--work-root', default='.',
                        help='Root directory (default: current dir)')
    parser.add_argument('--output-dir', default=None,
                        help='Output directory (default: work_dirs/paper_experiments/)')
    parser.add_argument('--tables', nargs='+',
                        default=['table2', 'table3', 'table4'],
                        help='Which tables to collect (default: all)')
    args = parser.parse_args()

    os.chdir(args.work_root)

    output_dir = args.output_dir or 'work_dirs/paper_experiments'
    os.makedirs(output_dir, exist_ok=True)

    write_fn = write_excel if HAS_OPENPYXL else write_csv_fallback
    if not HAS_OPENPYXL:
        print('[WARN] openpyxl not installed. Will output CSV instead of Excel.')
        print('       Install with: pip install openpyxl')

    table_configs = {
        'table2': ('Table2_ComponentAblation', TABLE2_EXPERIMENTS),
        'table3': ('Table3_MoE_Hyperparameters', TABLE3_EXPERIMENTS),
        'table4': ('Table4_SingleModal_vs_MultiModal', TABLE4_EXPERIMENTS),
    }

    for table_key in args.tables:
        if table_key not in table_configs:
            print(f'[WARN] Unknown table: {table_key}, skipping')
            continue

        table_name, experiments = table_configs[table_key]
        print(f'\n{"="*60}')
        print(f'Collecting {table_name}')
        print(f'{"="*60}')

        results = collect_table_data(experiments)

        ext = '.xlsx' if HAS_OPENPYXL else '.csv'
        output_path = os.path.join(output_dir, f'{table_name}{ext}')
        write_fn(results, experiments, output_path, table_name)

    print(f'\nDone! Results saved to {output_dir}/')


if __name__ == '__main__':
    main()
